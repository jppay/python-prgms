'''
Created on Jun 20, 2017

@author: Coalesce
'''
from openpyxl import Workbook, load_workbook 
from openpyxl.chart import (AreaChart, Reference, Series, BarChart3D)

wb = load_workbook('C:/Users/Coalesce.Coalesce-PC/Documents/Temp/LinkData.xlsx', read_only=False, data_only=True)
dataSheet = wb.get_sheet_by_name("Sheet1")

outList = []

# <cell_name_for_row>, <sheet_name>.cell(row=<cell_name_for_row>, column=col).value : gives cell nos and value
for eachCell in range(1, dataSheet.max_row): # For including all the rows with data
    listInside = []
    for col in range(3,5):                   # Columns have values
        listInside.append(dataSheet.cell(row=eachCell, column=col).value) # Method to get specific values 
    outList.append(listInside)

#print(outList)
#for i in outList: #Writes directly to the sheet
#    dataSheet.append(i)

chart = BarChart3D()
chart.title = "Link Utilization Graph"
chart.style = 13
chart.x_axis.title = 'Time'
chart.y_axis.title = 'Utilization (%)'

#Reference()
cats = Reference(dataSheet, min_col=1, min_row=2, max_col=2, max_row=17)
data = Reference(dataSheet, min_col=1, min_row=2, max_col=2, max_row=17)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
dataSheet.add_chart(chart, "F12")

wb.save('C:/Users/Coalesce.Coalesce-PC/Documents/Temp/LinkData.xlsx')

