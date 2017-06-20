'''
Created on Jun 20, 2017

@author: Coalesce
'''
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

fileName = 'C:/Users/Coalesce.Coalesce-PC/Documents/Temp/OpenPyXL.xlsx'
workbook1 = Workbook()
workbook1 = load_workbook(fileName, read_only=False) # Existing workbook loaded
sheetNames = workbook1.get_sheet_names() # Sheets in workbook checked

tempSheet = workbook1.get_sheet_by_name(sheetNames[11]) # Pick specific sheet (December)

# row, column are keywords to access each Cell
for ro in range(1,5):
    for col in range(5,10):
        tempSheet.cell(column=col, row=ro, value="%s" %(get_column_letter(col)) ) 

workbook1.save(fileName)
workbook1.close()